import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
import openpyxl
import os
import requests
import win32com.client

class ExcelMapperApp:
    def __init__(self, root):
        self.root = root
        self.icon = tk.PhotoImage(file = 'automation.png')
        self.root.iconphoto(False, self.icon)
        self.root.title("Excel Veri Aktarım Aracı")

        self.source_file = ""
        self.target_file = ""

        self.initialize_files()

        self.log_messages = []

        self.setup_ui()

    def setup_ui(self):
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)

        tk.Button(self.root, text="Kaynak Dosya Seç", command=self.select_source_file).grid(
            row=0, column=0, padx=5, pady=5, sticky="ew"
        )
        self.source_label = tk.Label(self.root, text="Seçilmedi")
        self.source_label.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        tk.Button(self.root, text="Hedef Dosya Seç", command=self.select_target_file).grid(
            row=1, column=0, padx=5, pady=5, sticky="ew"
        )
        self.target_label = tk.Label(self.root, text="Seçilmedi")
        self.target_label.grid(row=1, column=1, padx=5, pady=5, sticky="w")

        tk.Label(self.root, text="Hücre Eşleşmeleri (örn: A1 -> B1 veya A1,A2 -> C1)").grid(
            row=2, column=0, columnspan=2, pady=(10, 0)
        )

        self.mapping_text = scrolledtext.ScrolledText(self.root, width=60, height=15)
        self.mapping_text.grid(row=3, column=0, columnspan=2, padx=10, pady=5)

        tk.Button(self.root, text="Biçimlendir (-> Ekle)", command=self.format_arrows).grid(
            row=4, column=0, columnspan=2, pady=5
        )

        tk.Button(self.root, text="Eşleşmeleri Kaydet (.txt)", command=self.save_mappings).grid(
            row=5, column=0, padx=10, pady=5, sticky="ew"
        )
        tk.Button(self.root, text="Eşleşmeleri Yükle (.txt)", command=self.load_mappings).grid(
            row=5, column=1, padx=10, pady=5, sticky="ew"
        )

        tk.Button(self.root, text="Logları Göster", command=self.show_logs).grid(
            row=6, column=0, columnspan=2, padx=10, pady=5, sticky="ew"
        )

        tk.Button(self.root, text="Aktarımı Başlat", command=self.transfer_data).grid(
            row=7, column=0, columnspan=2, padx=10, pady=10, sticky="ew"
        )

        tk.Button(self.root, text="Raporları Çek", command=self.raporlari_cek).grid(
            row=8, column=0, padx=10, pady=5, sticky="ew"
        )

        self.kaynak_ekle_button = tk.Button(self.root, text="Kaynak Ekle", command=self.kaynak_penceresi_ac)
        self.kaynak_ekle_button.grid(
            row=8, column=1, padx=10, pady=5, sticky="ew"
        )

        self.info_text = tk.Text(self.root, height=4, width=50, state='disabled', bg="#f0f0f0")
        self.info_text.grid(row=9, column=0, columnspan=2, padx=10, pady=5)
    
    def initialize_files(self):
        if not os.path.exists("log.txt"):
            with open("log.txt", "w", encoding="utf-8") as f:
                f.write("Log dosyası oluşturuldu.\n")

        if not os.path.exists("kaynaklar.txt"):
            with open("kaynaklar.txt", "w", encoding="utf-8") as f:
                f.write("# Şirket ID'lerini buraya girin\n")


    def kaynak_penceresi_ac(self):
        pencere = tk.Toplevel(self.root)
        pencere.title("Kaynakları Görüntüle / Düzenle")
        pencere.geometry("400x300")

        pencere.grid_rowconfigure(0, weight=1)
        pencere.grid_columnconfigure(0, weight=1)

        frame = tk.Frame(pencere)
        frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        kaynak_text = tk.Text(frame, wrap="word")
        kaynak_text.grid(row=0, column=0, sticky="nsew")

        scrollbar = tk.Scrollbar(frame, orient="vertical", command=kaynak_text.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        kaynak_text.config(yscrollcommand=scrollbar.set)

        if os.path.exists("kaynaklar.txt"):
            with open("kaynaklar.txt", "r", encoding="utf-8") as f:
                kaynak_text.insert("1.0", f.read())

        def kaynak_kaydet():
            icerik = kaynak_text.get("1.0", tk.END).strip()
            with open("kaynaklar.txt", "w", encoding="utf-8") as f:
                f.write(icerik + "\n")
            self.log("Kaynaklar güncellendi.")
            messagebox.showinfo("Başarılı", "Kaynaklar başarıyla kaydedildi.")
            pencere.destroy()

        kaydet_buton = tk.Button(pencere, text="Kaydet", command=kaynak_kaydet)
        kaydet_buton.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))



    def log(self, message):
        self.log_messages.append(message)

    def show_logs(self):
        log_window = tk.Toplevel(self.root)
        log_window.title("Loglar")
        log_text = tk.Text(log_window, height=20, width=80)
        log_text.pack()
        log_text.insert(tk.END, "\n".join(self.log_messages))

    def select_source_file(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.source_file = file
            self.source_label.config(text=os.path.basename(file))
            self.show_source_info()
            self.log(f"Kaynak dosya seçildi: {file}")


    def show_source_info(self):
        wb = openpyxl.load_workbook(self.source_file)
        ws = wb.active
        val1 = ws["A1"].value
        val5 = ws["A5"].value
        val7 = ws["A7"].value
        info = f"A1: {val1}\nA5: {val5}\nA7: {val7}"
        self.info_text.config(state='normal')
        self.info_text.delete("1.0", tk.END)
        self.info_text.insert(tk.END, info)
        self.info_text.config(state='disabled')

    def select_target_file(self):
        file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file:
            self.target_file = file
            self.target_label.config(text=os.path.basename(file))
            self.log(f"Hedef dosya seçildi: {file}")

    def format_arrows(self):
        content = self.mapping_text.get("1.0", tk.END).strip().split("\n")
        new_lines = []
        for line in content:
            if "->" not in line and len(line.split()) == 2:
                left, right = line.split()
                new_lines.append(f"{left} -> {right}")
            else:
                new_lines.append(line)
        self.mapping_text.delete("1.0", tk.END)
        self.mapping_text.insert("1.0", "\n".join(new_lines))
        self.log("Eşleşmeler formatlandı.")

    def save_mappings(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
        if file_path:
            content = self.mapping_text.get("1.0", tk.END).strip()
            with open(file_path, "w", encoding="utf-8") as f:
                f.write(content)
            messagebox.showinfo("Başarılı", "Eşleşmeler kaydedildi.")
            self.log(f"Eşleşmeler kaydedildi: {file_path}")

    def load_mappings(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file_path:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()
            self.mapping_text.delete("1.0", tk.END)
            self.mapping_text.insert("1.0", content)
            self.log(f"Eşleşmeler yüklendi: {file_path}")

    def transfer_data(self):
        if not self.source_file or not self.target_file:
            messagebox.showerror("Hata", "Lütfen kaynak ve hedef dosyaları seçin.")
            return

        source_wb = openpyxl.load_workbook(self.source_file)
        source_ws = source_wb.active
        target_wb = openpyxl.load_workbook(self.target_file)
        target_ws = target_wb.active

        lines = self.mapping_text.get("1.0", tk.END).strip().split("\n")
        for line in lines:
            if "->" in line:
                left, right = map(str.strip, line.split("->"))
                source_cells = [cell.strip() for cell in left.split(",")]
                total = 0
                combined_text = ""
                numeric = True
                for cell in source_cells:
                    value = source_ws[cell].value
                    if isinstance(value, (int, float)):
                        total += value
                    else:
                        numeric = False
                        combined_text += str(value) + " "
                final_value = total if numeric else combined_text.strip()
                target_ws[right] = final_value


        target_wb.save(self.target_file)
        messagebox.showinfo("Tamamlandı", "Veriler başarıyla aktarıldı.")
        self.log("Transfer Işlemi Tamamlandı.")
        

    def raporlari_cek(self):
        # Mutlak yol kullanarak Reports klasörünü oluştur
        save_folder = os.path.abspath('./Reports')
        os.makedirs(save_folder, exist_ok=True)

        company_ids = []

        with open('kaynaklar.txt', 'r') as f:
            for line in list(f.readlines()):
                if line not in company_ids:
                    company_ids.append(str(line))

        for company_id in company_ids:
            print(f"\n📦 Şirket işleniyor: {company_id}")
            
            try:
                # 1. Excel URL'lerini al
                url = f'https://www.kap.org.tr/tr/api/company-detail/sgbf-data/{company_id}/FR/365'
                response = requests.get(url)
                disclosures = response.json()
                real_disclosures = []
                self.log(f'Şirket Bildirimleri Alınıyor : {company_id}')
                # Uygun raporları filtrele
                for disclosure in disclosures:
                    if disclosure['disclosureBasic']['title'] == 'Faaliyet Raporu (Konsolide Olmayan)' or disclosure['disclosureBasic']['title'] == 'Finansal Rapor':
                        if disclosure not in real_disclosures:
                            real_disclosures.append(disclosure)

                # Bildirim IDs'lerini topla
                notification_ids = []
                for disclosure in real_disclosures:
                    if str(disclosure['disclosureBasic']['disclosureIndex']) not in notification_ids:
                        notification_ids.append(disclosure['disclosureBasic']['disclosureIndex'])

                # Excel URL'lerini al
                excel_urls = []
                for _id in notification_ids:
                    excel_url = f'https://www.kap.org.tr/tr/api/notification/export/excel/{_id}'
                    if excel_url not in excel_urls:
                        excel_urls.append(excel_url)

                self.log(f'Exceller Toplandı : {company_id}')

                if not excel_urls:
                    print(f"🔍 {company_id} için uygun bildirim bulunamadı.")
                    continue

                # 2. Excel dosyalarını indir
                for url in excel_urls:
                    not_id = url.split('/')[-1]
                    file_name = f'Bildirim_{not_id}.xls'
                    save_path = os.path.join(save_folder, file_name)

                    headers = {
                    "User-Agent": "Mozilla/5.0"  # Gerekirse User-Agent ekleyin
                    }

                    try:
                        response = requests.get(url, headers=headers)

                        if response.status_code == 200:
                            with open(save_path, 'wb') as file:
                                file.write(response.content)
                            print(f'Excel Indirildi: {not_id}')
                            self.log(f'Excel Indirildi : {not_id}')
                        else:
                            print(f'Excel Indirilemedi: {not_id}')
                            self.log(f'Excel Indirilemedi : {not_id}')

                    except Exception as e:
                        print(f'{e}')

                    # 3. İndirilen .xls dosyalarını .xlsx'e çevir
                    xls_path = os.path.join(save_folder, f"Bildirim_{not_id}.xls")
                    xlsx_path = os.path.join(save_folder, f"Bildirim_{not_id}.xlsx")

                    if os.path.exists(xls_path):
                        try:
                            excel = win32com.client.Dispatch("Excel.Application")
                            workbook = excel.Workbooks.Open(xls_path)
                            workbook.SaveAs(xlsx_path, 51)  # 51, xlOpenXMLWorkbook türü (.xlsx)
                            workbook.Close(False)
                            excel.Quit()
                            print(f"'{xls_path}' başarıyla '{xlsx_path}' olarak kaydedildi.")
                            self.log(f'{not_id} başarıyla Excel olarak kaydedildi.')

                            # Başarıyla dönüştürülmüşse .xls dosyasını sil
                            os.remove(xls_path)
                            print(f"🗑️ Silindi: {xls_path}")
                            self.log(f'Dosyanın .xls formatı silindi : {not_id}')
                        except Exception as e:
                            print(f"Excel otomasyonunda bir hata oluştu: {e}")
                    else:
                        print(f"❗ Dosya bulunamadı: {xls_path}")
                        self.log(f'Dosya Bulunamadı : {xls_path}')

            except Exception as e:
                print(f"🚨 Hata oluştu ({company_id}): {e}")





if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMapperApp(root)
    root.mainloop()